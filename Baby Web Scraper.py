
# Author: Demiana Farid Michael
# Version: 1.0

"""
 Program: I have used 'openpyxl' library and 'beautifulsoup4' library to make an app that will show some statistics
          about a country the user enters. The statistics include the name of a state / province / governance and its population.
          It also has an option to display the state / province / governorate with the highest population and the one with the lowest population.

          The input is just the name of the country, or an abbreviation of it if the abbreviation is very popular.
"""

"""
Instructions :
1- Download openpyxl library
2- Download Beautifulsoup library
"""

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

while True:
    # Extracting data from the web
    cities = []  # cities list
    pop = []  # populations list


    def get_pop(req):
        temp = []  # creates temporary list to extract for modification
        total = 0
        if req.status_code == 200:
            html = req.text
            soup = BeautifulSoup(html, 'html.parser')

            # extract all text within the td tag in tbody
            for row in soup.find('tbody').find_all('tr'):
                temp.append([])
                for col in row.find_all('td'):
                    temp[len(temp) - 1].append(col.text)

            # row will be like ['','city', ... , 'Area', 'earliest population census',..., 'latest population census', '?', '?']
            for row in temp:
                cities.append(row[1])

            for row in temp:
                idx = len(row) - 3
                row[idx] = row[idx].replace(',', '')
                # iterate over population census from latest to earliest until it finds a number
                while idx > 5 and not row[idx].isnumeric():
                    idx -= 1
                    row[idx] = row[idx].replace(',', '')
                if row[idx - 1].replace(',', '').isnumeric():
                    pop.append(row[idx])
                else:
                    pop.append(0)

        else:
            print("Failed to get html.", req.status_code)
            print("Please check your spelling.")
            return False
        return True


    country = input("Enter country: ").lower()
    if country == 'israel':
        country = 'palestine'
        print('Did you mean Palestine?')
    country = country.replace(' ', '')

    url = "https://www.citypopulation.de/en/" + country + "/cities/"
    try:
        r = requests.get(url)
    except:
        print("Please check your internet connection.")
        continue
    try:
        if not get_pop(r):
            continue
    except:
        print("Sorry! The data is from a third party and it seems it's not complete. Please try another country.");
        continue

    # creating virtual worksheet

    wb = Workbook()
    ws = wb.active

    population = {}
    try:
        for i in range(len(cities)):
            population[cities[i]] = pop[i]
    except:
        print("Sorry! The data is from a third party and it seems it's not complete. Please try another country.");
        continue

    # filling the worksheet
    row = 1
    total = 0
    for key, value in population.items():
        if value == 0:
            continue
        row += 1
        # write data in cell 'A2' and 'B2' then 'A3' and 'B3' and so on
        ws['A' + str(row)] = key
        value = int(value)
        ws['B' + str(row)] = value
        population[key] = value
        total += value
    # write total of the country in cell 'A1' and 'B1'
    ws['A1'] = country.capitalize()
    # use SUM() function in excel
    ws['B1'] = "=SUM(B2:B" + str(row) + ")"
    try:
        wb.save(country.capitalize() + '.xlsx')
    except:
        print("Please close " + country.capitalize() + '.xlsx')
        continue

    print("Data is loaded successfully!")

    while True:
        # display options
        print("A: Display the population of each state / province / governorate and total population of the country")
        print(
            "B: Display the state / province / governorate with the highest population and the one with the lowest population.")
        print("C: Exit")
        option = input()
        if option not in ['a', 'b', 'c', 'A', 'B', 'C']:
            print("Invalid input. Please try again.")
            continue
        break

    if option == 'A' or option == 'a':
        # display all data in console
        print(country.capitalize(), total)
        for key, value in population.items():
            if value == 0:
                continue
            print(key, value)

    elif option == 'B' or option == 'b':
        # display the state / province / governorate with the highest population and the one with the lowest population
        max_pop = max(population.values())
        min_pop = min(population.values())
        for key, value in population.items():
            if value == max_pop:
                print('The state / province / governorate with the highest population is: ', key, value)
            if value == min_pop:
                print('The state / province / governorate with the lowest population is: ', key, value)
    else:
        break
